#!/usr/bin/env python3
"""
Extract VLOOKUP tables from Excel pricing workbook and convert to JSON.
This script extracts 7 pricing factor tables and 1 property type mapping table.
"""

import json
import os
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any

# Configuration
EXCEL_FILE = r"reference\directory\Base Pricing27.1_Pro_SMART_RCGV.xlsx"
OUTPUT_DIR = r"src\data\lookups"
SHEET_NAME = "VLOOKUP Tables"

def extract_two_column_table(ws, start_col, value_col, start_row=4, key_name="threshold", value_name="factor"):
    """Extract a two-column lookup table."""
    data = []
    row = start_row

    while True:
        key_cell = ws[f"{start_col}{row}"]
        value_cell = ws[f"{value_col}{row}"]

        # Stop if both cells are empty or None
        if key_cell.value is None and value_cell.value is None:
            break

        # Skip if either cell is empty (but continue looking)
        if key_cell.value is None or value_cell.value is None:
            row += 1
            if row > 100:  # Safety limit
                break
            continue

        # Convert to proper types
        key_val = key_cell.value
        value_val = value_cell.value

        # Handle numeric values - keep strings if they contain special chars like "+" or are already strings
        if isinstance(key_val, (int, float)):
            key_val = float(key_val) if '.' in str(key_val) or isinstance(key_val, float) else int(key_val)
        elif isinstance(key_val, str):
            # Keep as string if it has special characters like "55000+" or "6+"
            pass

        if isinstance(value_val, (int, float)):
            value_val = float(value_val)
        elif isinstance(value_val, str):
            # Keep as string if it has special characters
            pass

        data.append({
            key_name: key_val,
            value_name: value_val
        })

        row += 1
        if row > 100:  # Safety limit
            break

    return data

def extract_single_column_table(ws, col, start_row=4, value_name="factor"):
    """Extract a single-column table (for Multiple Properties)."""
    data = []
    row = start_row

    while True:
        cell = ws[f"{col}{row}"]

        if cell.value is None:
            break

        value_val = cell.value

        # Keep strings as-is (like "6+"), convert numbers to proper type
        if isinstance(value_val, (int, float)):
            value_val = int(value_val) if isinstance(value_val, int) or value_val == int(value_val) else float(value_val)

        # Index represents the number of properties (1-based)
        data.append({
            "propertyCount": value_val,  # Use the actual cell value (1, 2, 3, ... "6+")
            value_name: value_val  # Store the value itself for now (will need lookup in practice)
        })

        row += 1
        if row > 100:  # Safety limit
            break

    return data

def extract_property_type_mapping(ws, start_row=21, end_row=30):
    """Extract the property type mapping table (M21:O30)."""
    data = []

    for row in range(start_row, end_row + 1):
        property_type = ws[f"M{row}"].value
        code = ws[f"N{row}"].value
        depreciation_method = ws[f"O{row}"].value

        if property_type is None:
            continue

        # Convert numeric values
        if isinstance(code, (int, float)):
            code = int(code)
        if isinstance(depreciation_method, (int, float)):
            depreciation_method = int(depreciation_method)

        data.append({
            "propertyType": property_type,
            "code": code,
            "depreciationMethod": depreciation_method
        })

    return data

def main():
    """Main extraction function."""
    print(f"Loading Excel file: {EXCEL_FILE}")

    # Load workbook
    wb = load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found in workbook!")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    print(f"Successfully loaded sheet: {SHEET_NAME}")

    # Create output directory
    output_path = Path(OUTPUT_DIR)
    output_path.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {output_path.absolute()}")

    # Extract tables
    tables = {}

    print("\nExtracting tables...")

    # 1. Cost Basis Factor (Columns A-B)
    print("  - Cost Basis Factor (A-B)")
    tables['cost_basis'] = extract_two_column_table(
        ws, 'A', 'B', start_row=4,
        key_name="purchasePrice", value_name="factor"
    )

    # 2. Zip Code Factor (Columns D-E)
    print("  - Zip Code Factor (D-E)")
    tables['zip_code'] = extract_two_column_table(
        ws, 'D', 'E', start_row=4,
        key_name="zipCode", value_name="factor"
    )

    # 3. SqFt Factor (Columns G-H)
    print("  - SqFt Factor (G-H)")
    tables['sqft'] = extract_two_column_table(
        ws, 'G', 'H', start_row=4,
        key_name="squareFeet", value_name="factor"
    )

    # 4. Acres Factor (Columns J-K)
    print("  - Acres Factor (J-K)")
    tables['acres'] = extract_two_column_table(
        ws, 'J', 'K', start_row=4,
        key_name="acres", value_name="factor"
    )

    # 5. Property Type Factor (Columns M-N)
    print("  - Property Type Factor (M-N)")
    tables['property_type'] = extract_two_column_table(
        ws, 'M', 'N', start_row=4,
        key_name="propertyType", value_name="factor"
    )

    # 6. Number of Floors Factor (Columns P-Q)
    print("  - Number of Floors Factor (P-Q)")
    tables['floors'] = extract_two_column_table(
        ws, 'P', 'Q', start_row=4,
        key_name="numberOfFloors", value_name="factor"
    )

    # 7. Multiple Properties Factor (Columns S-T)
    print("  - Multiple Properties Factor (S-T)")
    tables['multiple_properties'] = extract_two_column_table(
        ws, 'S', 'T', start_row=4,
        key_name="propertyCount", value_name="factor"
    )

    # 8. Property Type Mapping (M21:O30)
    print("  - Property Type Mapping (M21:O30)")
    tables['property_type_mapping'] = extract_property_type_mapping(ws)

    # Save JSON files
    print("\nSaving JSON files...")

    file_mapping = {
        'cost_basis': 'cost-basis-factors.json',
        'zip_code': 'zip-code-factors.json',
        'sqft': 'sqft-factors.json',
        'acres': 'acres-factors.json',
        'property_type': 'property-type-factors.json',
        'floors': 'floor-factors.json',
        'multiple_properties': 'multiple-properties-factors.json',
        'property_type_mapping': 'property-types.json'
    }

    results = {}

    for key, filename in file_mapping.items():
        filepath = output_path / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(tables[key], f, indent=2, ensure_ascii=False)

        print(f"  [OK] {filename} ({len(tables[key])} records)")
        results[filename] = {
            'path': str(filepath.absolute()),
            'records': len(tables[key]),
            'sample': tables[key][:3] if len(tables[key]) > 0 else []
        }

    # Save a combined file for reference
    combined_path = output_path / 'all-lookup-tables.json'
    with open(combined_path, 'w', encoding='utf-8') as f:
        json.dump(tables, f, indent=2, ensure_ascii=False)
    print(f"  [OK] all-lookup-tables.json (combined reference)")

    # Generate summary report
    print("\n" + "="*60)
    print("EXTRACTION SUMMARY")
    print("="*60)

    for filename, info in results.items():
        print(f"\n{filename}:")
        print(f"  Location: {info['path']}")
        print(f"  Records: {info['records']}")
        if info['sample']:
            print(f"  Sample: {json.dumps(info['sample'][0], indent=4)}")

    print("\n" + "="*60)
    print("Extraction complete!")
    print("="*60)

    return results

if __name__ == "__main__":
    main()
