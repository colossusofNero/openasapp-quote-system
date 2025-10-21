import openpyxl
import json
from openpyxl.utils import get_column_letter

# Load the workbook
wb = openpyxl.load_workbook('reference/directory/Base Pricing27.1_Pro_SMART_RCGV.xlsx', data_only=False)

print("=" * 80)
print("EXCEL WORKBOOK ANALYSIS")
print("=" * 80)
print(f"\nWorkbook: Base Pricing27.1_Pro_SMART_RCGV.xlsx")
print(f"Number of sheets: {len(wb.sheetnames)}")
print(f"\nSheet names:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

print("\n" + "=" * 80)
print("DETAILED SHEET ANALYSIS")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    print(f"Dimensions: {max_row} rows x {max_col} columns")

    # Get merged cells
    merged = list(ws.merged_cells.ranges)
    if merged:
        print(f"Merged cells: {len(merged)} ranges")

    # Sample first 30 rows with data
    print(f"\nFirst 30 rows of data:")
    print("-" * 80)

    for row_idx in range(1, min(31, max_row + 1)):
        row_data = []
        has_data = False
        for col_idx in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns
            cell = ws.cell(row_idx, col_idx)
            value = cell.value

            # Check if cell has a formula
            if hasattr(cell, 'data_type') and cell.data_type == 'f':
                formula = f"={cell.value}" if cell.value else ""
                row_data.append(f"[FORMULA: {formula[:50]}...]" if len(formula) > 50 else f"[FORMULA: {formula}]")
                has_data = True
            elif value is not None:
                # Format the value
                if isinstance(value, (int, float)):
                    row_data.append(f"{value}")
                else:
                    row_data.append(f"{value}")
                has_data = True
            else:
                row_data.append("")

        if has_data:
            col_letters = [get_column_letter(i) for i in range(1, min(max_col + 1, 20))]
            print(f"Row {row_idx}: {dict(zip(col_letters, row_data))}")

    print("\n" + "-" * 80)

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
