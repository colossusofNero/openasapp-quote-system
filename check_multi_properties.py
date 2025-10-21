#!/usr/bin/env python3
"""Check the structure of Multiple Properties table."""

from openpyxl import load_workbook

EXCEL_FILE = r"reference\directory\Base Pricing27.1_Pro_SMART_RCGV.xlsx"
SHEET_NAME = "VLOOKUP Tables"

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

print("Multiple Properties Area (Columns R-U):")
print("-"*80)
for row in range(1, 20):
    r_val = ws[f"R{row}"].value
    s_val = ws[f"S{row}"].value
    t_val = ws[f"T{row}"].value
    u_val = ws[f"U{row}"].value

    if r_val or s_val or t_val or u_val:
        print(f"Row {row:2d}: R={r_val}, S={s_val}, T={t_val}, U={u_val}")
