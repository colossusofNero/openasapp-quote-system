#!/usr/bin/env python3
"""Check the property type mapping table structure."""

from openpyxl import load_workbook

EXCEL_FILE = r"reference\directory\Base Pricing27.1_Pro_SMART_RCGV.xlsx"
SHEET_NAME = "VLOOKUP Tables"

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

print("Property Type Mapping (M21:O30):")
print("-"*60)
for row in range(21, 31):
    m_val = ws[f"M{row}"].value
    n_val = ws[f"N{row}"].value
    o_val = ws[f"O{row}"].value
    print(f"Row {row}: M={m_val}, N={n_val}, O={o_val}")
