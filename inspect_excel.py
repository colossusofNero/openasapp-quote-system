#!/usr/bin/env python3
"""
Inspect the VLOOKUP Tables sheet to understand its structure.
"""

from openpyxl import load_workbook

EXCEL_FILE = r"reference\directory\Base Pricing27.1_Pro_SMART_RCGV.xlsx"
SHEET_NAME = "VLOOKUP Tables"

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

print("="*80)
print(f"Inspecting sheet: {SHEET_NAME}")
print("="*80)

# Inspect first 30 rows of relevant columns
columns_to_check = ['A', 'B', 'D', 'E', 'G', 'H', 'J', 'K', 'M', 'N', 'P', 'Q', 'S']

print("\nFirst 30 rows of data:")
print("-"*80)

for row in range(1, 31):
    row_data = []
    for col in columns_to_check:
        cell = ws[f"{col}{row}"]
        value = cell.value
        if value is not None:
            row_data.append(f"{col}{row}={value}")

    if row_data:
        print(f"Row {row:2d}: {', '.join(row_data)}")

print("\n" + "="*80)
print("Detailed inspection of each table area:")
print("="*80)

# Check Cost Basis Factor (A-B)
print("\nCost Basis Factor (Columns A-B):")
for row in range(1, 21):
    a_val = ws[f"A{row}"].value
    b_val = ws[f"B{row}"].value
    if a_val or b_val:
        print(f"  A{row}={a_val}, B{row}={b_val}")

# Check Zip Code Factor (D-E)
print("\nZip Code Factor (Columns D-E):")
for row in range(1, 21):
    d_val = ws[f"D{row}"].value
    e_val = ws[f"E{row}"].value
    if d_val or e_val:
        print(f"  D{row}={d_val}, E{row}={e_val}")

# Check SqFt Factor (G-H)
print("\nSqFt Factor (Columns G-H):")
for row in range(1, 21):
    g_val = ws[f"G{row}"].value
    h_val = ws[f"H{row}"].value
    if g_val or h_val:
        print(f"  G{row}={g_val}, H{row}={h_val}")

# Check Multiple Properties (S)
print("\nMultiple Properties Factor (Column S):")
for row in range(1, 21):
    s_val = ws[f"S{row}"].value
    if s_val:
        print(f"  S{row}={s_val}")
