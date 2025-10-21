import openpyxl
import json
from pathlib import Path

# Load the Excel workbook
excel_path = Path("reference/directory/Base Pricing27.1_Pro_SMART_RCGV.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=False)

# Extract formulas from specific sheets
formula_data = {
    "equation_sheet": {},
    "input_sheet": {},
    "vlookup_tables": {}
}

# Get Equation Sheet formulas
if "Equation Sheet" in wb.sheetnames:
    eq_sheet = wb["Equation Sheet"]

    # Key cells mentioned in the analysis
    key_cells = ["C21", "C51", "AB20", "AB26", "S4"]

    for cell_ref in key_cells:
        cell = eq_sheet[cell_ref]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            formula_data["equation_sheet"][cell_ref] = {
                "formula": cell.value,
                "value": cell.value
            }

    # Scan entire sheet for formulas
    for row in eq_sheet.iter_rows(min_row=1, max_row=60, min_col=1, max_col=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                coord = cell.coordinate
                if coord not in formula_data["equation_sheet"]:
                    formula_data["equation_sheet"][coord] = {
                        "formula": cell.value,
                        "row": cell.row,
                        "col": cell.column
                    }

# Get Input Sheet formulas
if "Input Sheet" in wb.sheetnames:
    input_sheet = wb["Input Sheet"]

    # Key output cells
    key_cells = ["B9", "B14", "B15", "B17", "D17", "F17", "D11", "D14"]

    for cell_ref in key_cells:
        cell = input_sheet[cell_ref]
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_data["input_sheet"][cell_ref] = {
                    "formula": cell.value,
                    "description": f"Cell {cell_ref}"
                }
            else:
                formula_data["input_sheet"][cell_ref] = {
                    "value": str(cell.value),
                    "type": type(cell.value).__name__
                }

    # Scan for more formulas
    for row in input_sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=30):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                coord = cell.coordinate
                if coord not in formula_data["input_sheet"]:
                    formula_data["input_sheet"][coord] = {
                        "formula": cell.value,
                        "row": cell.row,
                        "col": cell.column
                    }

# Get VLOOKUP Tables structure
if "VLOOKUP Tables" in wb.sheetnames:
    vlookup_sheet = wb["VLOOKUP Tables"]

    # Extract factor tables
    tables = {
        "cost_basis": {"start_col": "A", "factor_col": "B", "rows": range(2, 20)},
        "zip_code": {"start_col": "D", "factor_col": "E", "rows": range(2, 20)},
        "sqft": {"start_col": "G", "factor_col": "H", "rows": range(2, 20)},
        "acres": {"start_col": "J", "factor_col": "K", "rows": range(2, 20)},
        "property_type": {"start_col": "M", "factor_col": "N", "rows": range(2, 15)},
        "floors": {"start_col": "P", "factor_col": "Q", "rows": range(2, 15)},
    }

    for table_name, table_info in tables.items():
        formula_data["vlookup_tables"][table_name] = []
        for row_num in table_info["rows"]:
            key_cell = vlookup_sheet[f"{table_info['start_col']}{row_num}"]
            factor_cell = vlookup_sheet[f"{table_info['factor_col']}{row_num}"]

            if key_cell.value is not None:
                formula_data["vlookup_tables"][table_name].append({
                    "key": key_cell.value,
                    "factor": factor_cell.value
                })

# Save to JSON for analysis
output_path = Path("formula_analysis.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(formula_data, f, indent=2, default=str)

print(f"Formula extraction complete! Saved to {output_path}")
print(f"\nEquation Sheet formulas found: {len(formula_data['equation_sheet'])}")
print(f"Input Sheet formulas found: {len(formula_data['input_sheet'])}")
print(f"VLOOKUP tables extracted: {len(formula_data['vlookup_tables'])}")

# Print key formulas
print("\n=== KEY FORMULAS ===")
print("\nEquation Sheet:")
for cell, data in list(formula_data["equation_sheet"].items())[:10]:
    print(f"  {cell}: {data.get('formula', 'N/A')[:80]}...")

print("\nInput Sheet:")
for cell, data in list(formula_data["input_sheet"].items())[:10]:
    formula = data.get('formula', data.get('value', 'N/A'))
    print(f"  {cell}: {str(formula)[:80]}...")
